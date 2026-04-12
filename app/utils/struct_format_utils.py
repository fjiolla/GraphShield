import io
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import magic
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

struct_logger = logging.getLogger("struct_format_utils")

# ─────────────────────────────────────────────
# MIME Type Constants
# ─────────────────────────────────────────────
STRUCT_MIME_CSV = "text/csv"
STRUCT_MIME_JSON = "application/json"
STRUCT_MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
STRUCT_MIME_XLS = "application/vnd.ms-excel"
STRUCT_MIME_SQL_MARKERS = ["text/plain", "application/sql", "text/x-sql"]

STRUCT_SUPPORTED_MIMES = {
    "csv": [STRUCT_MIME_CSV, "text/comma-separated-values"],
    "json": [STRUCT_MIME_JSON, "text/json"],
    "xlsx": [STRUCT_MIME_XLSX, STRUCT_MIME_XLS],
    "sql": STRUCT_MIME_SQL_MARKERS,
}


# ─────────────────────────────────────────────
# MIME Detection
# ─────────────────────────────────────────────
def struct_detect_mime(file_path: str | Path) -> str:
    """
    Detect MIME type of a file using python-magic (libmagic).

    Args:
        file_path: Absolute or relative path to the file.

    Returns:
        MIME type string (e.g., 'text/csv').

    Raises:
        FileNotFoundError: If the file does not exist.
        RuntimeError: If MIME detection fails.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    try:
        mime = magic.from_file(str(file_path), mime=True)
        struct_logger.info("Detected MIME '%s' for file: %s", mime, file_path.name)
        return mime
    except Exception as exc:
        raise RuntimeError(f"MIME detection failed for {file_path}: {exc}") from exc


# def struct_resolve_format(mime_type: str, file_extension: str = "") -> str:
#     """
#     Resolve a canonical format string ('csv', 'json', 'xlsx', 'sql')
#     from a MIME type and optional file extension fallback.

#     Args:
#         mime_type: MIME type string.
#         file_extension: File extension without dot (e.g., 'csv').

#     Returns:
#         Format string: 'csv' | 'json' | 'xlsx' | 'sql' | 'unknown'.
#     """
#     mime_lower = mime_type.lower()

#     for fmt, mimes in STRUCT_SUPPORTED_MIMES.items():
#         if mime_lower in [m.lower() for m in mimes]:
#             return fmt

#     # SQL files are often detected as text/plain — use extension as tiebreaker
#     if mime_lower in STRUCT_MIME_SQL_MARKERS and file_extension.lower() == "sql":
#         return "sql"

#     ext_map = {"csv": "csv", "json": "json", "xlsx": "xlsx", "xls": "xlsx", "sql": "sql"}
#     if file_extension.lower() in ext_map:
#         struct_logger.warning(
#             "MIME '%s' unrecognized; falling back to extension '%s'.",
#             mime_type,
#             file_extension,
#         )
#         return ext_map[file_extension.lower()]

#     struct_logger.error("Cannot resolve format from MIME '%s' or extension '%s'.", mime_type, file_extension)
#     return "unknown"
def struct_resolve_format(mime_type: str, file_extension: str = "") -> str:
    mime_lower = mime_type.lower()
    ext = file_extension.lower()

    # ✅ STEP 1: PRIORITY → extension (FIXED)
    ext_map = {
        "csv": "csv",
        "json": "json",
        "xlsx": "xlsx",
        "xls": "xlsx",
        "sql": "sql"
    }

    if ext in ext_map:
        return ext_map[ext]

    # ✅ STEP 2: fallback → MIME
    for fmt, mimes in STRUCT_SUPPORTED_MIMES.items():
        if mime_lower in [m.lower() for m in mimes]:
            return fmt

    # ✅ STEP 3: SQL fallback
    if mime_lower in STRUCT_MIME_SQL_MARKERS:
        return "sql"

    struct_logger.error(
        "Cannot resolve format from MIME '%s' or extension '%s'.",
        mime_type,
        file_extension,
    )
    return "unknown"


# ─────────────────────────────────────────────
# Encoding Detection
# ─────────────────────────────────────────────
def struct_detect_encoding(file_path: str | Path) -> str:
    """
    Auto-detect file encoding by reading the first 64 KB.

    Args:
        file_path: Path to the file.

    Returns:
        Encoding string (e.g., 'utf-8', 'latin-1').
    """
    try:
        import chardet

        with open(file_path, "rb") as fh:
            raw = fh.read(65536)
        result = chardet.detect(raw)
        encoding = result.get("encoding") or "utf-8"
        struct_logger.info("Detected encoding '%s' (confidence %.2f).", encoding, result.get("confidence", 0))
        return encoding
    except ImportError:
        struct_logger.warning("chardet not installed; defaulting to utf-8.")
        return "utf-8"
    except Exception as exc:
        struct_logger.warning("Encoding detection failed (%s); defaulting to utf-8.", exc)
        return "utf-8"


# ─────────────────────────────────────────────
# JSON Flattening
# ─────────────────────────────────────────────
def struct_flatten_json(
    data: dict | list,
    parent_key: str = "",
    sep: str = ".",
    _depth: int = 0,
    _max_depth: int = 10,
) -> dict:
    """
    Recursively flatten a nested dict/list into dot-notation keys.

    Example:
        {"user": {"address": {"city": "Mumbai"}}}
        → {"user.address.city": "Mumbai"}

    Args:
        data: Dict or list to flatten.
        parent_key: Prefix for nested keys.
        sep: Separator character.
        _depth: Internal recursion depth tracker.
        _max_depth: Maximum recursion depth (guard against deeply nested docs).

    Returns:
        Flat dictionary with dot-notation keys.
    """
    items: dict[str, Any] = {}

    if _depth > _max_depth:
        struct_logger.warning("Max JSON flatten depth reached at key '%s'. Storing as string.", parent_key)
        return {parent_key: str(data)}

    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else str(key)
            if isinstance(value, (dict, list)):
                items.update(
                    struct_flatten_json(value, new_key, sep, _depth + 1, _max_depth)
                )
            else:
                items[new_key] = value

    elif isinstance(data, list):
        for idx, item in enumerate(data):
            new_key = f"{parent_key}{sep}{idx}" if parent_key else str(idx)
            if isinstance(item, (dict, list)):
                items.update(
                    struct_flatten_json(item, new_key, sep, _depth + 1, _max_depth)
                )
            else:
                items[new_key] = item
    else:
        items[parent_key] = data

    return items


def struct_json_to_dataframe(raw_json: list | dict) -> pd.DataFrame:
    """
    Convert a JSON object (list of records or nested dict) to a flat DataFrame.

    Args:
        raw_json: Parsed JSON (list of dicts or a single dict).

    Returns:
        pandas DataFrame with flattened columns.

    Raises:
        ValueError: If the JSON cannot be converted to a tabular structure.
    """
    if isinstance(raw_json, dict):
        # Try to find a list-of-records within the dict
        for val in raw_json.values():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                raw_json = val
                break
        else:
            raw_json = [raw_json]

    if not isinstance(raw_json, list):
        raise ValueError("JSON root must be a list of records or a dict.")

    flat_records = [struct_flatten_json(record) for record in raw_json]
    df = pd.DataFrame(flat_records)
    struct_logger.info("JSON flattened to DataFrame: %d rows × %d cols.", len(df), len(df.columns))
    return df


# ─────────────────────────────────────────────
# Excel Utilities
# ─────────────────────────────────────────────
def struct_forward_fill_merged_cells(ws) -> None:
    """
    Forward-fill values in merged cells so every cell has an explicit value.
    Modifies the openpyxl worksheet in place.

    Args:
        ws: openpyxl Worksheet object.
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for merge_range in merged_ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        fill_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=merge_range.max_row,
            min_col=min_col,
            max_col=merge_range.max_col,
        ):
            for cell in row:
                if cell.value is None:
                    cell.value = fill_value


def struct_detect_header_row(ws, max_check_rows: int = 10) -> int:
    """
    Detect the header row index in an Excel sheet using density analysis
    (the row with the highest proportion of non-empty string cells).

    Args:
        ws: openpyxl Worksheet object.
        max_check_rows: Number of rows from the top to examine.

    Returns:
        0-based row index of the most likely header row.
    """
    best_row = 0
    best_score = -1.0

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_check_rows)):
        if not row:
            continue
        string_count = sum(1 for cell in row if isinstance(cell.value, str) and cell.value.strip())
        total = len(row)
        score = string_count / total if total > 0 else 0.0
        if score > best_score:
            best_score = score
            best_row = row_idx

    struct_logger.info("Detected header at row index %d (score=%.2f).", best_row, best_score)
    return best_row


_EXCEL_EPOCH = datetime(1899, 12, 30)


def struct_convert_excel_date(value: Any) -> Any:
    """
    Convert an Excel serial date (float/int) or datetime to ISO 8601 string.

    Args:
        value: Cell value from openpyxl.

    Returns:
        ISO 8601 date string if the value is date-like, else the original value.
    """
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            dt = _EXCEL_EPOCH + pd.Timedelta(days=float(value))
            return dt.isoformat()
        except Exception:
            pass
    return value


def struct_xlsx_to_dataframe(file_path: str | Path) -> pd.DataFrame:
    """
    Load an Excel file into a pandas DataFrame.
    Handles:
      - Merged cell forward-filling
      - Automatic header row detection (first 10 rows)
      - Excel serial date → ISO 8601 conversion
      - Mixed data types

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        pandas DataFrame.

    Raises:
        ValueError: If the file cannot be parsed as Excel.
    """
    file_path = Path(file_path)
    try:
        wb = load_workbook(str(file_path), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file '{file_path}': {exc}") from exc

    # Forward-fill merged cells
    struct_forward_fill_merged_cells(ws)

    # Detect header row
    header_row_idx = struct_detect_header_row(ws)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")

    headers = [
        str(cell).strip() if cell is not None else f"col_{i}"
        for i, cell in enumerate(rows[header_row_idx])
    ]

    # Deduplicate header names
    seen: dict[str, int] = {}
    deduped_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped_headers.append(h)

    data_rows = rows[header_row_idx + 1 :]
    records = []
    for row in data_rows:
        record = {
            deduped_headers[i]: struct_convert_excel_date(val)
            for i, val in enumerate(row)
            if i < len(deduped_headers)
        }
        records.append(record)

    df = pd.DataFrame(records, columns=deduped_headers)
    df.dropna(how="all", inplace=True)
    struct_logger.info("Excel loaded: %d rows × %d cols.", len(df), len(df.columns))
    return df


# ─────────────────────────────────────────────
# DataFrame Sanitization
# ─────────────────────────────────────────────
def struct_sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean a DataFrame before writing to SQLite:
    - Strip whitespace from string columns
    - Coerce columns with mixed numeric/string to string
    - Drop fully-empty columns
    - Normalize column names to snake_case

    Args:
        df: Input DataFrame.

    Returns:
        Sanitized DataFrame.
    """
    # Normalize column names
    df.columns = [
        re.sub(r"[^\w]", "_", str(c)).strip("_").lower() or f"col_{i}"
        for i, c in enumerate(df.columns)
    ]

    # Drop fully-empty columns
    df.dropna(axis=1, how="all", inplace=True)

    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)

            # Detect mixed-type columns: try numeric conversion
            numeric_attempt = pd.to_numeric(df[col], errors="coerce")
            null_ratio = numeric_attempt.isna().mean()
            original_null_ratio = df[col].isna().mean()

            if null_ratio > original_null_ratio + 0.3:
                # More than 30% new nulls introduced → mixed type → keep as string
                df[col] = df[col].astype(str).replace("nan", None)

    struct_logger.info("DataFrame sanitized: %d rows × %d cols.", len(df), len(df.columns))
    return df


def struct_safe_table_name(name: str) -> str:
    """
    Convert a string into a safe SQLite table name.

    Args:
        name: Raw name (e.g., filename without extension).

    Returns:
        Sanitized table name.
    """
    safe = re.sub(r"[^\w]", "_", name).strip("_").lower()
    safe = re.sub(r"_+", "_", safe)
    if safe and safe[0].isdigit():
        safe = f"t_{safe}"
    return safe or "struct_table"